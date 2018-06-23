import click
import MySQLdb
from openpyxl import load_workbook
import os
import django
os.environ['DJANGO_SETTINGS_MODULE']='summerMRND.settings'
django.setup()
from onlineapp.models import *

USERNAME = "root"
PASSWORD ="root"
HOST ="localhost"
DBNAME="mrndsummer"

file1='new.xlsx'
file2='mock.xlsx'

def already_exists(dbname):
    conn = MySQLdb.connect(HOST, USERNAME, PASSWORD)
    c = conn.cursor()
    return True if c.execute("SHOW DATABASES LIKE '%s';"%(dbname)) else False
@click.group()
def sqli():
    pass
@sqli.command()
def createdb():
    """This command can createa the database in the MYSQL server and if the data base already exists , this prints message"""
    dbname=DBNAME
    if not already_exists(dbname):
        conn = MySQLdb.connect(HOST, USERNAME, PASSWORD)
        c = conn.cursor()
        c.execute("CREATE DATABASE IF NOT EXISTS %s;"%(dbname))
        c.execute("USE %s"%(dbname))
        c.execute("CREATE TABLE students(name varchar(50), college varchar(10),emailId varchar(50),DBname varchar(20));")
        c.execute("CREATE TABLE marks(sno int,student varchar(50),question1 int , question2 int,question3 int, question4 int, total int);")
        c.close()
        click.echo("Database sucessfullly created with tables students and marks")
    else:
        click.echo("Database already exists!")

@sqli.command()
def dropdb():
    """This is to delete the database from the MySQL server"""
    dbname=DBNAME
    if click.confirm("Are you sure you wish to frop the database?") and already_exists(dbname):
        conn = MySQLdb.connect(HOST, USERNAME, PASSWORD)
        c = conn.cursor()
        c.execute("DROP DATABASE %s;"%(dbname))
        c.close()
        click.echo("dropped sucessfully!")
def get_required_sheet(file,sheet):
    try:
        wb=load_workbook(file)
    except:
        click.echo("Reacheck the file names!")
    if sheet in wb.sheetnames:
        return wb[sheet]
    return None
def getValue(sheet,row,col):
    value=sheet.cell(row=row,column=col).value
    return int(value) if value.strip().isdigit() else value

@sqli.command()
def importdb():
    """this command is to import the required data in the database"""
    sheet=get_required_sheet(file1,'Colleges')
    for row in range(2,sheet.max_row+1):
        c=College(name=getValue(sheet,row,1) ,location=getValue(sheet,row,3),acronym=getValue(sheet,row,2),contact=getValue(sheet,row,4) )
        c.save()
    click.echo('college data inserted')
    sheet=get_required_sheet(file1,"Current")
    x = College.objects.order_by('-location')
    for row in range(2,sheet.max_row+1):
        try:
            c=Student(name=getValue(sheet, row, 1), college=College.objects.get(acronym=getValue(sheet,row,2)),email=getValue(sheet,row,3),db_folder=getValue(sheet,row,4).lower())
            c.save()
        except:
            print(f"values for this college not found {getValue(sheet,row,2)}")
    #Student.objects.get(db_folder__contains=)
    click.echo("student data inserted")
    sheet=get_required_sheet(file2,'Sheet')
    for row in range(2,sheet.max_row+1):
        try:
            c=MockTest1(student=Student.objects.get(db_folder=getValue(sheet,row,2).split('_')[-2]),problem1=getValue(sheet,row,3),problem2=getValue(sheet,row,4),problem3=getValue(sheet,row,5),problem4=getValue(sheet,row,6),total=getValue(sheet,row,7))
            c.save()
        except:
            print(f"values for this student not found {getValue(sheet,row,2)}")
    click.echo("marks data inserted")
    sheet=get_required_sheet(file1,'Deletions')
    for row in range(2,sheet.max_row+1):
        try:
            c = Student(name=getValue(sheet, row, 1), college=College.objects.get(acronym=getValue(sheet, row, 2)),
                        email=getValue(sheet, row, 3), db_folder=getValue(sheet, row, 4).lower(),dropped_out=True)
            c.save()
        except:
            print(f"values for this college not found {getValue(sheet,row,2)}")
    click.echo("Dropped out imported to db")

@sqli.command()
def clearData():
    """clears all the data from the data base"""
    Student.objects.all().delete()
    College.objects.all().delete()
    MockTest1.objects.all().delete()
    click.echo("all data deleted")
@sqli.command()
def test():
    """just for testing purpouses"""
    click.echo("working")

if __name__=="__main__":
    sqli()